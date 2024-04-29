from openpyxl import Workbook, load_workbook
import os
import HeadImage



# 输入文件夹和输出 excel 文件路径
input_folder = 'images/13'
output_file = 'data/head13.xlsx'

# 创建或打开目标 Excel 文件
if os.path.exists(output_file):
    workbook = load_workbook(filename=output_file)
else:
    workbook = Workbook()

# 获取当前工作表
sheet = workbook.active
sheet.title = "Result"


# 添加列表头
header = ['PicName','label','pitch', 'yaw', 'roll']
sheet.append(header)

# 设置固定列宽
for column_letter in ["A", "B", "C", "D", "E"]:
    column_dimension = sheet.column_dimensions[column_letter]
    column_dimension.width =15

# 遍历图片文件夹，计算参数并保存到 Excel 中
for file_name in os.listdir(input_folder):

    if file_name.endswith(".jpg"):
        image_path = os.path.join(input_folder, file_name)
        print(image_path)
        pitch,yaw,roll=HeadImage.HeadInfo(image_path)
        print(pitch,yaw,roll)
        label = file_name.split('_')[0].strip()
        # 将文件名和计算结果添加到 Excel 表格中
        row_data = [file_name,label,pitch,yaw,roll]
        sheet.append(row_data)

# 保存 Excel 文件
workbook.save(filename=output_file)
