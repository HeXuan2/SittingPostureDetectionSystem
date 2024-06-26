from openpyxl import Workbook, load_workbook
import os
import cv2
import CrossLegImage

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Alignment

# 输入文件夹和输出 excel 文件路径
input_folder = 'images/51/5/'
output_file = 'data/CrossLeg51_5.xlsx'

# 创建或打开目标 Excel 文件
if os.path.exists(output_file):
    workbook = load_workbook(filename=output_file)
else:
    workbook = Workbook()

# 获取当前工作表
sheet = workbook.active
sheet.title = "Result"



# 添加列表头
header = ['PicName','label','distance1', 'distance2', 'distance3', 'distance4','distance5', 'distance6']
sheet.append(header)

# 设置固定列宽
for column_letter in ["A", "B", "C", "D", "E","F","G","H"]:
    column_dimension = sheet.column_dimensions[column_letter]
    column_dimension.width =15

# 遍历图片文件夹，计算参数并保存到 Excel 中
for file_name in os.listdir(input_folder):
    if file_name.endswith(".jpg"):
        image_path = os.path.join(input_folder, file_name)
        img = cv2.imread(image_path)

        print(image_path)
        crossDot1, distance1, crossDot2, distance2, crossDot3, distance3, distance4, crossDot5, distance5, distance6 =\
            CrossLegImage.CrossLegImageInfo(image_path)

        label = file_name.split('_')[0].strip()
        # 将文件名和计算结果添加到 Excel 表格中
        row_data = [file_name,label,distance1, distance2, distance3, distance4,distance5,distance6]
        sheet.append(row_data)


# 保存 Excel 文件
workbook.save(filename=output_file)
